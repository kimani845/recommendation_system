import pandas as pd
import numpy as np
import matplotlib.pyplot as plt
import seaborn as sns
from sklearn.model_selection import train_test_split
from sklearn.ensemble import RandomForestRegressor
from sklearn.metrics import mean_squared_error, mean_absolute_error
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.chart import BarChart, Reference, LineChart
from datetime import datetime, timedelta
import calendar
import os

# # Constants
# CAKE_TYPES = ['Heart Cakes', 'Coconut Cakes', 'Mobile Cakes', 'Block Cakes','Star Cakes', 'Queen Cakes', 'Sweet Cakes']
# REGIONS = ['Whitehouse', 'Ngomongo', 'Kiamunyi', 'Kabachia'] 
# DAYS_OF_WEEK = ['Monday', 'Tuesday', 'Wednesday', 'Thursday', 'Friday', 'Saturday', 'Sunday']

class CakeSalesTracker:
    def __init__(self):
        pass # Removed the excel sheet that i was using before
        
        # Create the Excel file if it doesn't exist
        if not os.path.exists(excel_file):
            self.create_excel_structure()
    
    def create_excel_structure(self):
        """Create the initial Excel file structure with all necessary sheets"""
        workbook = openpyxl.Workbook()
        
        # Remove default sheet
        default_sheet = workbook.active
        workbook.remove(default_sheet)
        
        # Create Daily Sales sheet
        daily_sheet = workbook.create_sheet('Daily Sales')
        
        # Create header row
        headers = ['Date', 'Day of Week', 'Region']
        headers.extend(CAKE_TYPES)
        headers.append('Total Sales')
        
        for col, header in enumerate(headers, 1):
            cell = daily_sheet.cell(row=1, column=col)
            cell.value = header
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="DDEBF7", end_color="DDEBF7", fill_type="solid")
        
        # Create Weekly Summary sheet
        weekly_sheet = workbook.create_sheet('Weekly Summary')
        weekly_headers = ['Week', 'Start Date', 'End Date'] + CAKE_TYPES + ['Total Sales']
        
        for col, header in enumerate(weekly_headers, 1):
            cell = weekly_sheet.cell(row=1, column=col)
            cell.value = header
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="DDEBF7", end_color="DDEBF7", fill_type="solid")
        
        # Create Monthly Analysis sheet
        monthly_sheet = workbook.create_sheet('Monthly Analysis')
        monthly_headers = ['Month', 'Year'] + CAKE_TYPES + ['Total Sales']
        
        for col, header in enumerate(monthly_headers, 1):
            cell = monthly_sheet.cell(row=1, column=col)
            cell.value = header
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="DDEBF7", end_color="DDEBF7", fill_type="solid")
        
        # Create Day of Week Analysis sheet
        dow_sheet = workbook.create_sheet('Day of Week Analysis')
        dow_headers = ['Day of Week'] + CAKE_TYPES + ['Total Sales']
        
        for col, header in enumerate(dow_headers, 1):
            cell = dow_sheet.cell(row=1, column=col)
            cell.value = header
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="DDEBF7", end_color="DDEBF7", fill_type="solid")
        
        # Add days of week
        for row, day in enumerate(DAYS_OF_WEEK, 2):
            dow_sheet.cell(row=row, column=1).value = day
        
        # Create Regional Analysis sheet
        region_sheet = workbook.create_sheet('Regional Analysis')
        region_headers = ['Region'] + CAKE_TYPES + ['Total Sales']
        
        for col, header in enumerate(region_headers, 1):
            cell = region_sheet.cell(row=1, column=col)
            cell.value = header
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="DDEBF7", end_color="DDEBF7", fill_type="solid")
        
        # Add regions
        for row, region in enumerate(REGIONS, 2):
            region_sheet.cell(row=row, column=1).value = region
        
        # Create Predictions sheet
        pred_sheet = workbook.create_sheet('Predictions')
        pred_headers = ['Date', 'Day of Week', 'Region'] + CAKE_TYPES
        
        for col, header in enumerate(pred_headers, 1):
            cell = pred_sheet.cell(row=1, column=col)
            cell.value = header
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="DDEBF7", end_color="DDEBF7", fill_type="solid")
        
        # Create Dashboard sheet
        dashboard = workbook.create_sheet('Dashboard')
        dashboard.cell(row=1, column=1).value = "Cake Sales Dashboard"
        dashboard.cell(row=1, column=1).font = Font(size=16, bold=True)
        
        # Save the workbook
        workbook.save(self.excel_file)
        print(f"Created Excel file: {self.excel_file}")
    
    def load_data(self):
        """Load data from Excel file into pandas DataFrame"""
        try:
            self.sales_data = pd.read_excel(self.excel_file, sheet_name='Daily Sales')
            return True
        except Exception as e:
            print(f"Error loading data: {e}")
            return False
    
    def add_daily_sales(self, date, region, sales_dict):
        """
        Add daily sales data to the Excel file
        
        Parameters:
        - date: Date of sales (string in format 'YYYY-MM-DD' or datetime object)
        - region: Region name (string)
        - sales_dict: Dictionary with cake types as keys and quantities as values
        """
        if isinstance(date, str):
            date = datetime.strptime(date, '%Y-%m-%d')
        
        day_of_week = DAYS_OF_WEEK[date.weekday()]
        
        # Load the workbook
        workbook = openpyxl.load_workbook(self.excel_file)
        daily_sheet = workbook['Daily Sales']
        
        # Find the next empty row
        next_row = daily_sheet.max_row + 1
        
        # Add the date, day of week, and region
        daily_sheet.cell(row=next_row, column=1).value = date
        daily_sheet.cell(row=next_row, column=2).value = day_of_week
        daily_sheet.cell(row=next_row, column=3).value = region
        
        # Add sales for each cake type
        total_sales = 0
        for col, cake_type in enumerate(CAKE_TYPES, 4):
            sales = sales_dict.get(cake_type, 0)
            daily_sheet.cell(row=next_row, column=col).value = sales
            total_sales += sales
        
        # Add total sales
        daily_sheet.cell(row=next_row, column=len(CAKE_TYPES) + 4).value = total_sales
        
        # Save the workbook
        workbook.save(self.excel_file)
        print(f"Added sales data for {date.strftime('%Y-%m-%d')}")
    
    def update_summaries(self):
        """Update weekly, monthly, and analysis sheets based on daily sales data"""
        if not self.load_data():
            print("No data to summarize")
            return
        
        # Convert date column to datetime if it's not already
        if not pd.api.types.is_datetime64_any_dtype(self.sales_data['Date']):
            self.sales_data['Date'] = pd.to_datetime(self.sales_data['Date'])
        
        # Create weekly summary
        self.sales_data['Week'] = self.sales_data['Date'].dt.isocalendar().week
        self.sales_data['Year'] = self.sales_data['Date'].dt.isocalendar().year
        
        weekly_summary = self.sales_data.groupby(['Year', 'Week']).agg({
            **{cake: 'sum' for cake in CAKE_TYPES},
            'Total Sales': 'sum'
        }).reset_index()
        
        # Add start and end date for each week
        weekly_summary['Start Date'] = weekly_summary.apply(
            lambda row: datetime.fromisocalendar(int(row['Year']), int(row['Week']), 1), axis=1
        )
        weekly_summary['End Date'] = weekly_summary.apply(
            lambda row: datetime.fromisocalendar(int(row['Year']), int(row['Week']), 7), axis=1
        )
        
        # Reorder columns
        weekly_summary = weekly_summary[['Week', 'Start Date', 'End Date'] + CAKE_TYPES + ['Total Sales']]
        
        # Create monthly summary
        self.sales_data['Month'] = self.sales_data['Date'].dt.month
        monthly_summary = self.sales_data.groupby(['Year', 'Month']).agg({
            **{cake: 'sum' for cake in CAKE_TYPES},
            'Total Sales': 'sum'
        }).reset_index()
        
        # Create day of week summary
        dow_summary = self.sales_data.groupby('Day of Week').agg({
            **{cake: 'sum' for cake in CAKE_TYPES},
            'Total Sales': 'sum'
        }).reset_index()
        
        # Create region summary
        region_summary = self.sales_data.groupby('Region').agg({
            **{cake: 'sum' for cake in CAKE_TYPES},
            'Total Sales': 'sum'
        }).reset_index()
        
        # Update Excel sheets
        with pd.ExcelWriter(self.excel_file, engine='openpyxl', mode='a', if_sheet_exists='replace') as writer:
            weekly_summary.to_excel(writer, sheet_name='Weekly Summary', index=False)
            monthly_summary.to_excel(writer, sheet_name='Monthly Analysis', index=False)
            dow_summary.to_excel(writer, sheet_name='Day of Week Analysis', index=False)
            region_summary.to_excel(writer, sheet_name='Regional Analysis', index=False)
        
        print("Updated summary sheets")
    
    def train_prediction_model(self):
        """Train a machine learning model to predict sales"""
        if not self.load_data():
            print("No data to train model")
            return
        
        # Convert date column to datetime if it's not already
        if not pd.api.types.is_datetime64_any_dtype(self.sales_data['Date']):
            self.sales_data['Date'] = pd.to_datetime(self.sales_data['Date'])
        
        # Create features
        self.sales_data['DayOfWeek'] = self.sales_data['Date'].dt.dayofweek
        self.sales_data['Month'] = self.sales_data['Date'].dt.month
        self.sales_data['DayOfMonth'] = self.sales_data['Date'].dt.day
        
        # One-hot encode region
        region_dummies = pd.get_dummies(self.sales_data['Region'], prefix='Region')
        self.sales_data = pd.concat([self.sales_data, region_dummies], axis=1)
        
        # Train a model for each cake type
        self.models = {}
        
        for cake_type in CAKE_TYPES:
            print(f"Training model for {cake_type}...")
            
            # Prepare features and target
            features = ['DayOfWeek', 'Month', 'DayOfMonth'] + [col for col in self.sales_data.columns if col.startswith('Region_')]
            X = self.sales_data[features]
            y = self.sales_data[cake_type]
            
            # Split data
            X_train, X_test, y_train, y_test = train_test_split(X, y, test_size=0.2, random_state=42)
            
            # Train model
            model = RandomForestRegressor(n_estimators=100, random_state=42)
            model.fit(X_train, y_train)
            
            # Evaluate model
            y_pred = model.predict(X_test)
            mse = mean_squared_error(y_test, y_pred)
            mae = mean_absolute_error(y_test, y_pred)
            
            print(f"  MSE: {mse:.2f}, MAE: {mae:.2f}")
            
            # Store model
            self.models[cake_type] = model
        
        print("Model training complete")
    
    def predict_next_day(self, date=None, region=None):
        """
        Predict sales for the next day
        
        Parameters:
        - date: Date to predict for (defaults to tomorrow)
        - region: Region to predict for (required)
        
        Returns:
        - Dictionary with predicted sales for each cake type
        """
        if not hasattr(self, 'models') or not self.models:
            print("No trained models available. Please train models first.")
            return None
        
        if date is None:
            date = datetime.now() + timedelta(days=1)
        elif isinstance(date, str):
            date = datetime.strptime(date, '%Y-%m-%d')
        
        if region is None:
            print("Region is required for prediction")
            return None
        
        # Create features for prediction
        features = {
            'DayOfWeek': date.weekday(),
            'Month': date.month,
            'DayOfMonth': date.day
        }
        
        # Add one-hot encoded region
        for r in REGIONS:
            features[f'Region_{r}'] = 1 if r == region else 0
        
        # Create DataFrame for prediction
        pred_df = pd.DataFrame([features])
        
        # Make predictions for each cake type
        predictions = {}
        for cake_type, model in self.models.items():
            pred = model.predict(pred_df)[0]
            predictions[cake_type] = max(0, round(pred))  # Ensure non-negative and round to nearest integer
        
        return predictions
    
    def add_prediction_to_excel(self, date, region, predictions):
        """
        Add prediction to the Excel file
        
        Parameters:
        - date: Date of prediction
        - region: Region name
        - predictions: Dictionary with cake types as keys and predicted quantities as values
        """
        if isinstance(date, str):
            date = datetime.strptime(date, '%Y-%m-%d')
        
        day_of_week = DAYS_OF_WEEK[date.weekday()]
        
        # Load the workbook
        workbook = openpyxl.load_workbook(self.excel_file)
        pred_sheet = workbook['Predictions']
        
        # Find the next empty row
        next_row = pred_sheet.max_row + 1
        
        # Add the date, day of week, and region
        pred_sheet.cell(row=next_row, column=1).value = date
        pred_sheet.cell(row=next_row, column=2).value = day_of_week
        pred_sheet.cell(row=next_row, column=3).value = region
        
        # Add predictions for each cake type
        for col, cake_type in enumerate(CAKE_TYPES, 4):
            pred_sheet.cell(row=next_row, column=col).value = predictions.get(cake_type, 0)
        
        # Save the workbook
        workbook.save(self.excel_file)
        print(f"Added prediction for {date.strftime('%Y-%m-%d')} in {region}")
    
    def update_dashboard(self):
        """Update the dashboard sheet with charts and key metrics"""
        if not self.load_data():
            print("No data to update dashboard")
            return
        
        workbook = openpyxl.load_workbook(self.excel_file)
        dashboard = workbook['Dashboard']
        
        # Clear existing content
        for row in dashboard.iter_rows(min_row=2):
            for cell in row:
                cell.value = None
        
        # Add title
        dashboard.cell(row=1, column=1).value = "Cake Sales Dashboard"
        dashboard.cell(row=1, column=1).font = Font(size=16, bold=True)
        
        # Add cake type sales chart
        dashboard.cell(row=3, column=1).value = "Total Sales by Cake Type"
        dashboard.cell(row=3, column=1).font = Font(bold=True)
        
        cake_sales = self.sales_data[CAKE_TYPES].sum().reset_index()
        cake_sales.columns = ['Cake Type', 'Total Sales']
        
        # Write data for chart
        for i, (cake, sales) in enumerate(zip(cake_sales['Cake Type'], cake_sales['Total Sales'])):
            dashboard.cell(row=4+i, column=1).value = cake
            dashboard.cell(row=4+i, column=2).value = sales
        
        # Create chart
        chart1 = BarChart()
        chart1.title = "Total Sales by Cake Type"
        chart1.y_axis.title = "Sales"
        chart1.x_axis.title = "Cake Type"
        
        data = Reference(dashboard, min_col=2, min_row=4, max_row=4+len(CAKE_TYPES)-1)
        cats = Reference(dashboard, min_col=1, min_row=4, max_row=4+len(CAKE_TYPES)-1)
        chart1.add_data(data)
        chart1.set_categories(cats)
        
        dashboard.add_chart(chart1, "D3")
        
        # Add day of week sales chart
        dashboard.cell(row=3, column=6).value = "Sales by Day of Week"
        dashboard.cell(row=3, column=6).font = Font(bold=True)
        
        dow_sales = self.sales_data.groupby('Day of Week')['Total Sales'].sum().reset_index()
        
        # Reorder days of week
        dow_order = {day: i for i, day in enumerate(DAYS_OF_WEEK)}
        dow_sales['order'] = dow_sales['Day of Week'].map(dow_order)
        dow_sales = dow_sales.sort_values('order').drop('order', axis=1)
        
        # Write data for chart
        for i, (day, sales) in enumerate(zip(dow_sales['Day of Week'], dow_sales['Total Sales'])):
            dashboard.cell(row=4+i, column=6).value = day
            dashboard.cell(row=4+i, column=7).value = sales
        
        # Create chart
        chart2 = BarChart()
        chart2.title = "Sales by Day of Week"
        chart2.y_axis.title = "Sales"
        chart2.x_axis.title = "Day of Week"
        
        data = Reference(dashboard, min_col=7, min_row=4, max_row=4+len(DAYS_OF_WEEK)-1)
        cats = Reference(dashboard, min_col=6, min_row=4, max_row=4+len(DAYS_OF_WEEK)-1)
        chart2.add_data(data)
        chart2.set_categories(cats)
        
        dashboard.add_chart(chart2, "I3")
        
        # Add regional sales chart
        dashboard.cell(row=15, column=1).value = "Sales by Region"
        dashboard.cell(row=15, column=1).font = Font(bold=True)
        
        region_sales = self.sales_data.groupby('Region')['Total Sales'].sum().reset_index()
        
        # Write data for chart
        for i, (region, sales) in enumerate(zip(region_sales['Region'], region_sales['Total Sales'])):
            dashboard.cell(row=16+i, column=1).value = region
            dashboard.cell(row=16+i, column=2).value = sales
        
        # Create chart
        chart3 = BarChart()
        chart3.title = "Sales by Region"
        chart3.y_axis.title = "Sales"
        chart3.x_axis.title = "Region"
        
        data = Reference(dashboard, min_col=2, min_row=16, max_row=16+len(REGIONS)-1)
        cats = Reference(dashboard, min_col=1, min_row=16, max_row=16+len(REGIONS)-1)
        chart3.add_data(data)
        chart3.set_categories(cats)
        
        dashboard.add_chart(chart3, "D15")
        
        # Add key metrics
        dashboard.cell(row=15, column=6).value = "Key Metrics"
        dashboard.cell(row=15, column=6).font = Font(bold=True)
        
        # Total sales
        dashboard.cell(row=16, column=6).value = "Total Sales:"
        dashboard.cell(row=16, column=7).value = self.sales_data['Total Sales'].sum()
        
        # Best selling cake
        best_cake = cake_sales.loc[cake_sales['Total Sales'].idxmax(), 'Cake Type']
        dashboard.cell(row=17, column=6).value = "Best Selling Cake:"
        dashboard.cell(row=17, column=7).value = best_cake
        
        # Best sales day
        best_day = dow_sales.loc[dow_sales['Total Sales'].idxmax(), 'Day of Week']
        dashboard.cell(row=18, column=6).value = "Best Sales Day:"
        dashboard.cell(row=18, column=7).value = best_day
        
        # Best region
        best_region = region_sales.loc[region_sales['Total Sales'].idxmax(), 'Region']
        dashboard.cell(row=19, column=6).value = "Best Region:"
        dashboard.cell(row=19, column=7).value = best_region
        
        # Save the workbook
        workbook.save(self.excel_file)
        print("Updated dashboard")

# Example usage
if __name__ == "__main__":
    tracker = CakeSalesTracker()
    
    # Example: Add sample data for April 2023
    import random
    
    # Generate sample data for April 2023
    start_date = datetime(2023, 4, 1)
    end_date = datetime(2023, 4, 30)
    
    current_date = start_date
    while current_date <= end_date:
        for region in REGIONS:
            # Generate random sales data
            sales = {cake: random.randint(10, 50) for cake in CAKE_TYPES}
            
            # Add some patterns:
            # - Weekends have higher sales
            if current_date.weekday() >= 5:  # Saturday or Sunday
                for cake in sales:
                    sales[cake] = int(sales[cake] * 1.5)
            
            # - Heart cakes sell better on Fridays
            if current_date.weekday() == 4:  # Friday
                sales['Heart Cakes'] = int(sales['Heart Cakes'] * 1.3)
            
            # - Region preferences
            if region == 'Whitehouse':
                sales['Coconut Cakes'] = int(sales['Coconut Cakes'] * 1.2)
            elif region == 'Ngomongo':
                sales['Mobile Cakes'] = int(sales['Mobile Cakes'] * 1.2)
            elif region == 'Kiamunyi':
                sales['Queen Cakes'] = int(sales['Queen Cakes'] * 1.2)
            elif region == 'Kabachia':
                sales['Sweet Cakes'] = int(sales['Sweet Cakes'] * 1.2)
            
            tracker.add_daily_sales(current_date, region, sales)
        
        current_date += timedelta(days=1)
    
    # Update summaries and dashboard
    tracker.update_summaries()
    tracker.update_dashboard()
    
    # Train prediction model
    tracker.train_prediction_model()
    
    # Make prediction for next day
    next_day = datetime(2023, 5, 1)
    for region in REGIONS:
        predictions = tracker.predict_next_day(next_day, region)
        if predictions:
            tracker.add_prediction_to_excel(next_day, region, predictions)
            print(f"Predictions for {next_day.strftime('%Y-%m-%d')} in {region}:")
            for cake, quantity in predictions.items():
                print(f"  {cake}: {quantity}")